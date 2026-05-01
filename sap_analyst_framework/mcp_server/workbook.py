"""
workbook.py — Low-level workbook manipulation helpers for the MCP server.

Uses openpyxl directly (no VBA, no win32com) so it can run headlessly on any OS.
All write operations preserve existing VBA macros when the file is .xlsm.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Theme colours (mirrors modTheme.bas) ─────────────────────────────────────
GOLD        = "CCAA66"
CHARCOAL    = "1F1F1F"
LIGHT_GRAY  = "FAFAFA"
WHITE       = "FFFFFF"
ERROR_RED   = "DC3545"
SUCCESS_GRN = "28A745"
MID_GRAY    = "6C757D"

# ── Sheet definitions ─────────────────────────────────────────────────────────
SHEET_HEADERS: dict[str, list[str]] = {
    "HOME":          [],
    "SETTINGS":      ["Key", "Value"],
    "REPORT_CONFIG": ["KPI_Name", "Formula_Type", "Column_Name", "Target_Sheet",
                      "Target_Cell", "Filter_Column", "Filter_Value", "Label",
                      "Filters", "Threshold_Green", "Threshold_Orange", "Threshold_Red",
                      "Format", "DependsOn", "DataSheet"],
    "UI_CONFIG":     ["Component", "Sheet", "Row", "Col", "Width", "Height",
                      "Label", "Value_Cell", "Style", "Color_Override"],
    "DATA":          [],
    "LOG":           ["Timestamp", "Step", "Status", "Message"],
    "HISTORY":       ["Timestamp", "Report", "Plant", "Date From", "Date To", "User",
                      "Duration (sec)", "Status"],
}

REQUIRED_SHEET_ORDER = ["HOME", "SETTINGS", "REPORT_CONFIG", "UI_CONFIG",
                         "DATA", "LOG", "HISTORY"]


# ── Workbook creation ─────────────────────────────────────────────────────────

def setup_workbook(path: str | Path, overwrite: bool = False) -> dict[str, Any]:
    """
    Create (or update) the framework workbook with all required sheets.

    If the file already exists:
    - overwrite=False : adds only missing sheets (safe to call multiple times)
    - overwrite=True  : recreates all sheets from scratch (destructive)

    Returns a summary dict.
    """
    p = Path(path)
    created_sheets: list[str] = []
    updated_sheets: list[str] = []

    if p.exists() and not overwrite:
        wb = load_workbook(p, keep_vba=True)
        existing = set(wb.sheetnames)
        for name in REQUIRED_SHEET_ORDER:
            if name not in existing:
                wb.create_sheet(name)
                _write_header(wb[name], SHEET_HEADERS.get(name, []))
                _style_header(wb[name])
                created_sheets.append(name)
    else:
        wb = Workbook()
        # Rename the default sheet to HOME
        wb.active.title = "HOME"
        for name in REQUIRED_SHEET_ORDER[1:]:
            wb.create_sheet(name)
        for name in REQUIRED_SHEET_ORDER:
            _write_header(wb[name], SHEET_HEADERS.get(name, []))
            _style_header(wb[name])
        created_sheets = REQUIRED_SHEET_ORDER[:]

    _apply_tab_colours(wb)
    _setup_home_sheet(wb["HOME"])

    wb.save(p)
    wb.close()

    return {
        "path":           str(p),
        "created_sheets": created_sheets,
        "updated_sheets": updated_sheets,
        "message":        f"Workbook ready at {p}",
    }


# ── Settings ──────────────────────────────────────────────────────────────────

def write_settings(path: str | Path, settings: dict[str, str]) -> dict[str, Any]:
    """Write key-value pairs to the SETTINGS sheet."""
    p = Path(path)
    wb = load_workbook(p, keep_vba=p.suffix == ".xlsm")
    ws = wb["SETTINGS"]

    # Build index of existing keys → row number
    key_index: dict[str, int] = {}
    for row in ws.iter_rows(min_row=2):
        k = str(row[0].value or "").strip()
        if k:
            key_index[k.lower()] = row[0].row

    for key, value in settings.items():
        if key.lower() in key_index:
            ws.cell(row=key_index[key.lower()], column=2, value=value)
        else:
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1, value=key)
            ws.cell(row=next_row, column=2, value=value)
            key_index[key.lower()] = next_row

    wb.save(p)
    wb.close()
    return {"written": list(settings.keys())}


def read_settings(path: str | Path) -> dict[str, str]:
    """Return all SETTINGS key-value pairs as a dict."""
    p = Path(path)
    wb = load_workbook(p, read_only=True, keep_vba=p.suffix == ".xlsm", data_only=True)
    ws = wb["SETTINGS"]
    result: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        k = str(row[0] or "").strip()
        v = str(row[1] or "").strip()
        if k:
            result[k] = v
    wb.close()
    return result


# ── REPORT_CONFIG ─────────────────────────────────────────────────────────────

def add_kpi(path: str | Path, kpi: dict[str, str]) -> dict[str, Any]:
    """Append one KPI row to REPORT_CONFIG."""
    p = Path(path)
    wb = load_workbook(p, keep_vba=p.suffix == ".xlsm")
    ws = wb["REPORT_CONFIG"]
    headers = SHEET_HEADERS["REPORT_CONFIG"]
    next_row = ws.max_row + 1
    for col_idx, header in enumerate(headers, start=1):
        key = header.lower().replace(" ", "_")
        ws.cell(row=next_row, column=col_idx, value=kpi.get(key, ""))
    wb.save(p)
    wb.close()
    return {"added_row": next_row, "kpi_name": kpi.get("kpi_name", "")}


def list_kpis(path: str | Path) -> list[dict[str, str]]:
    """Return all rows from REPORT_CONFIG as a list of dicts."""
    p = Path(path)
    wb = load_workbook(p, read_only=True, keep_vba=p.suffix == ".xlsm", data_only=True)
    ws = wb["REPORT_CONFIG"]
    headers = [str(c or "").strip() for c in next(ws.iter_rows(max_row=1, values_only=True), [])]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {headers[i]: str(row[i] or "") for i in range(len(headers))}
        if any(d.values()):
            result.append(d)
    wb.close()
    return result


def clear_kpis(path: str | Path) -> dict[str, Any]:
    """Remove all KPI rows from REPORT_CONFIG (keeps header)."""
    p = Path(path)
    wb = load_workbook(p, keep_vba=p.suffix == ".xlsm")
    ws = wb["REPORT_CONFIG"]
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
    # Delete empty rows
    for r in range(ws.max_row, 1, -1):
        if all(ws.cell(r, c).value is None for c in range(1, 9)):
            ws.delete_rows(r)
    wb.save(p)
    wb.close()
    return {"cleared": True}


# ── UI_CONFIG ─────────────────────────────────────────────────────────────────

def add_ui_component(path: str | Path, component: dict[str, Any]) -> dict[str, Any]:
    """Append one row to UI_CONFIG."""
    p = Path(path)
    wb = load_workbook(p, keep_vba=p.suffix == ".xlsm")
    ws = wb["UI_CONFIG"]
    headers = SHEET_HEADERS["UI_CONFIG"]
    next_row = ws.max_row + 1
    for col_idx, header in enumerate(headers, start=1):
        key = header.lower().replace(" ", "_").replace("-", "_")
        ws.cell(row=next_row, column=col_idx, value=component.get(key, ""))
    wb.save(p)
    wb.close()
    return {"added_row": next_row}


def clear_ui_config(path: str | Path) -> None:
    """Remove all UI_CONFIG rows (keeps header)."""
    p = Path(path)
    wb = load_workbook(p, keep_vba=p.suffix == ".xlsm")
    ws = wb["UI_CONFIG"]
    for r in range(ws.max_row, 1, -1):
        ws.delete_rows(r)
    wb.save(p)
    wb.close()


# ── High-level dashboard builder ──────────────────────────────────────────────

def build_kpi_dashboard(
    path: str | Path,
    report_name: str,
    tcode: str,
    date_mode: str,
    plant: str,
    variant: str,
    layout: str,
    kpis: list[dict[str, str]],
    use_python: bool = False,
    loader_type: str = "BAR",
) -> dict[str, Any]:
    """
    All-in-one: write SETTINGS + REPORT_CONFIG + UI_CONFIG for a full KPI dashboard.

    kpis: list of dicts with keys:
        kpi_name, formula_type, column_name, target_cell,
        filter_column (opt), filter_value (opt), label (opt)

    Automatically assigns target_sheet = HOME and calculates grid positions
    for KPI cards in AUTO mode.
    """
    p = Path(path)

    # Ensure workbook has all sheets
    setup_workbook(p, overwrite=False)

    # Write settings
    settings = {
        "ReportName":      report_name,
        "TCode":           tcode,
        "DateMode":        date_mode,
        "Plant":           plant,
        "Variant":         variant,
        "Layout":          layout,
        "UI_MODE":         "AUTO",
        "LOADER_TYPE":     loader_type,
        "UsePython":       "YES" if use_python else "NO",
        "ExportFileName":  f"{report_name.replace(' ', '_')}_Export",
    }
    write_settings(p, settings)

    # Write REPORT_CONFIG
    clear_kpis(p)
    kpi_cells: list[str] = []
    for i, kpi in enumerate(kpis):
        target_col = chr(ord("E") + (i % 8))  # E..L, cycles per row
        target_row = 8 + (i // 8) * 4
        cell = f"{target_col}{target_row}"
        kpi_cells.append(cell)
        add_kpi(p, {
            "kpi_name":    kpi.get("kpi_name", f"KPI_{i+1}"),
            "formula_type": kpi.get("formula_type", "SUM"),
            "column_name": kpi.get("column_name", ""),
            "target_sheet": "HOME",
            "target_cell": kpi.get("target_cell", cell),
            "filter_column": kpi.get("filter_column", ""),
            "filter_value": kpi.get("filter_value", ""),
            "label": kpi.get("label", kpi.get("kpi_name", "")),
        })

    # Write UI_CONFIG — section header + one KPI card per KPI + Run button
    clear_ui_config(p)
    wb = load_workbook(p, keep_vba=p.suffix == ".xlsm")
    ws_ui = wb["UI_CONFIG"]
    _write_header(ws_ui, SHEET_HEADERS["UI_CONFIG"])

    components: list[list[Any]] = []

    # Section header at row 3
    components.append(["SECTION_HEADER", "HOME", 3, 1, 8, 1, report_name, "", "", ""])

    # Run Report button at row 5
    components.append(["BUTTON", "HOME", 5, 1, 3, 1, "Run Report", "", "modMain.RunMain", ""])

    # KPI cards — 4 per row, starting at row 8
    cards_per_row = 4
    card_col_span = 2
    card_row_span = 3
    start_row = 8

    for i, kpi in enumerate(kpis):
        col_in_row = i % cards_per_row
        row_group  = i // cards_per_row
        card_row   = start_row + row_group * (card_row_span + 1)
        card_col   = 1 + col_in_row * (card_col_span + 1)
        cell       = kpi.get("target_cell", kpi_cells[i] if i < len(kpi_cells) else f"E{start_row}")
        components.append([
            "KPI_CARD", "HOME",
            card_row, card_col,
            card_col_span, card_row_span,
            kpi.get("label", kpi.get("kpi_name", f"KPI {i+1}")),
            f"HOME!{cell}",
            "", "",
        ])

    for c_idx, comp in enumerate(components, start=2):
        for col_idx, val in enumerate(comp, start=1):
            ws_ui.cell(row=c_idx, column=col_idx, value=val)

    wb.save(p)
    wb.close()

    return {
        "report_name": report_name,
        "path":        str(p),
        "kpi_count":   len(kpis),
        "ui_components": len(components),
        "message": (
            f"Dashboard configured: {len(kpis)} KPI(s), {len(components)} UI components. "
            f"Import VBA modules then click Run Report on HOME."
        ),
    }


# ── Internal helpers ──────────────────────────────────────────────────────────

def _write_header(ws, headers: list[str]) -> None:
    if not headers:
        return
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)


def _style_header(ws) -> None:
    """Apply Gold-on-Charcoal styling to the header row."""
    if ws.max_row < 1:
        return
    hdr_fill = PatternFill("solid", fgColor=CHARCOAL)
    hdr_font = Font(color=GOLD, bold=True, size=10)
    hdr_align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        if cell.value:
            cell.fill  = hdr_fill
            cell.font  = hdr_font
            cell.alignment = hdr_align


def _apply_tab_colours(wb: Workbook) -> None:
    """Set tab colour to Charcoal for all framework sheets."""
    for name in REQUIRED_SHEET_ORDER:
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = CHARCOAL


def _setup_home_sheet(ws) -> None:
    """Write a minimal HOME layout skeleton."""
    # Title bar row 1
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value     = "SAP Analyst Framework"
    title_cell.font      = Font(color=GOLD, bold=True, size=16)
    title_cell.fill      = PatternFill("solid", fgColor=CHARCOAL)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Status row 3
    ws["A3"].value = "Status:"
    ws["A3"].font  = Font(bold=True)
    ws["B3"].value = "Ready"

    # Progress row 4
    ws["A4"].value = "Progress:"
    ws["A4"].font  = Font(bold=True)
    ws["B4"].value = "0%"

    # Column widths
    ws.column_dimensions["A"].width = 18
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 14
