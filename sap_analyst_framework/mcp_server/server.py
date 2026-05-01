"""
server.py — MCP server for the SAP Analyst Excel Framework.

Exposes tools that Claude (in VS Code or Claude Code) can call to build,
configure, and inspect the Excel workbook without touching the VBE manually.

Usage:
    python server.py           # stdio transport (Claude Code / VS Code)
    python server.py --dev     # development mode with verbose logging

Registration (Claude Code):
    Add to .claude/settings.json → mcpServers (see docs/mcp_server.md)

Requirements:
    pip install mcp openpyxl
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

from mcp.server.fastmcp import FastMCP

# Ensure sibling modules are importable when launched from any cwd
sys.path.insert(0, str(Path(__file__).parent))
import workbook as wb_ops

# ── Server instance ───────────────────────────────────────────────────────────
mcp = FastMCP(
    name="sap-analyst-framework",
    instructions=(
        "Tools for building and configuring SAP Analyst Excel Framework workbooks. "
        "Use build_kpi_dashboard for one-shot dashboard creation. "
        "Use setup_workbook + write_settings + add_kpi for step-by-step setup."
    ),
)


# ── Tool: setup_workbook ──────────────────────────────────────────────────────
@mcp.tool()
def setup_workbook(
    path: str,
    overwrite: bool = False,
) -> str:
    """
    Create (or update) the SAP Analyst Framework workbook with all required sheets:
    HOME, SETTINGS, REPORT_CONFIG, UI_CONFIG, DATA, LOG, HISTORY.

    If the file already exists, only missing sheets are added (safe to re-run).
    Set overwrite=True to recreate all sheets from scratch.

    Args:
        path: Full path to the .xlsx or .xlsm file to create/update.
        overwrite: If True, recreate all sheets. Default False.
    """
    try:
        result = wb_ops.setup_workbook(path, overwrite=overwrite)
        return json.dumps(result, indent=2)
    except Exception as exc:
        return json.dumps({"error": str(exc)})


# ── Tool: write_settings ──────────────────────────────────────────────────────
@mcp.tool()
def write_settings(
    path: str,
    settings_json: str,
) -> str:
    """
    Write configuration settings to the SETTINGS sheet.

    Args:
        path: Path to the workbook.
        settings_json: JSON object mapping setting keys to values.
            Available keys: ProjectName, ReportName, Plant, TCode, Variant,
            Layout, DateMode, DateFrom, DateTo, ExportFileName, UI_MODE,
            LOADER_TYPE, UsePython, PythonPath, SAPConnectionName.

    Example settings_json:
        {"TCode": "MB51", "DateMode": "PREVIOUS_WEEK", "Plant": "1000"}
    """
    try:
        settings = json.loads(settings_json)
        result = wb_ops.write_settings(path, settings)
        return json.dumps(result)
    except Exception as exc:
        return json.dumps({"error": str(exc)})


# ── Tool: read_settings ───────────────────────────────────────────────────────
@mcp.tool()
def read_settings(path: str) -> str:
    """
    Read all current settings from the SETTINGS sheet.

    Args:
        path: Path to the workbook.
    """
    try:
        result = wb_ops.read_settings(path)
        return json.dumps(result, indent=2)
    except Exception as exc:
        return json.dumps({"error": str(exc)})


# ── Tool: add_kpi ─────────────────────────────────────────────────────────────
@mcp.tool()
def add_kpi(
    path: str,
    kpi_name: str,
    formula_type: str,
    column_name: str,
    target_cell: str,
    target_sheet: str = "HOME",
    filter_column: str = "",
    filter_value: str = "",
    label: str = "",
) -> str:
    """
    Add one KPI row to the REPORT_CONFIG sheet.

    Args:
        path:          Path to the workbook.
        kpi_name:      Descriptive name (e.g. "Total Withdrawn Qty").
        formula_type:  SUM | COUNT | COUNT_UNIQUE | AVERAGE | MIN | MAX | FORMULA | CUSTOM
        column_name:   Header name in the DATA sheet to aggregate (case-insensitive).
        target_cell:   Cell where the result is written (e.g. "E10").
        target_sheet:  Sheet containing target_cell. Default: HOME.
        filter_column: Optional column to filter by before aggregating.
        filter_value:  Value to match in filter_column (exact, case-insensitive).
        label:         Human-readable label shown on KPI cards.
    """
    try:
        result = wb_ops.add_kpi(path, {
            "kpi_name":     kpi_name,
            "formula_type": formula_type,
            "column_name":  column_name,
            "target_sheet": target_sheet,
            "target_cell":  target_cell,
            "filter_column": filter_column,
            "filter_value":  filter_value,
            "label":         label or kpi_name,
        })
        return json.dumps(result)
    except Exception as exc:
        return json.dumps({"error": str(exc)})


# ── Tool: list_kpis ───────────────────────────────────────────────────────────
@mcp.tool()
def list_kpis(path: str) -> str:
    """
    List all KPI definitions currently in REPORT_CONFIG.

    Args:
        path: Path to the workbook.
    """
    try:
        result = wb_ops.list_kpis(path)
        return json.dumps(result, indent=2)
    except Exception as exc:
        return json.dumps({"error": str(exc)})


# ── Tool: add_ui_component ────────────────────────────────────────────────────
@mcp.tool()
def add_ui_component(
    path: str,
    component: str,
    sheet: str,
    row: int,
    col: int,
    width: int = 2,
    height: int = 1,
    label: str = "",
    value_cell: str = "",
    style: str = "",
) -> str:
    """
    Add one component row to the UI_CONFIG sheet (used when UI_MODE=AUTO).

    Args:
        path:       Path to the workbook.
        component:  BUTTON | KPI_CARD | STATUS_BOX | SECTION_HEADER |
                    NAV_BUTTON | PROGRESS_BAR | PROGRESS_CIRCLE | ALERT_BOX
        sheet:      Target sheet name (e.g. "HOME").
        row:        Excel row number where the component is placed.
        col:        Excel column number (1=A, 2=B, …).
        width:      Column span. Default 2.
        height:     Row span. Default 1.
        label:      Text displayed in/on the component.
        value_cell: For KPI_CARD: formula cell reference e.g. "HOME!E10".
                    For NAV_BUTTON: target sheet name.
                    For BUTTON: macro name e.g. "modMain.RunMain".
        style:      For STATUS_BOX / ALERT_BOX: OK | ERROR | WARN.
    """
    try:
        result = wb_ops.add_ui_component(path, {
            "component":   component,
            "sheet":       sheet,
            "row":         row,
            "col":         col,
            "width":       width,
            "height":      height,
            "label":       label,
            "value_cell":  value_cell,
            "style":       style,
            "color_override": "",
        })
        return json.dumps(result)
    except Exception as exc:
        return json.dumps({"error": str(exc)})


# ── Tool: build_kpi_dashboard — the one-shot tool ─────────────────────────────
@mcp.tool()
def build_kpi_dashboard(
    path: str,
    report_name: str,
    tcode: str,
    date_mode: str,
    plant: str,
    kpis_json: str,
    variant: str = "",
    layout: str = "",
    use_python: bool = False,
    loader_type: str = "BAR",
) -> str:
    """
    ONE-SHOT: Configure the entire KPI dashboard from a single call.

    Creates the workbook (if needed), writes SETTINGS, populates REPORT_CONFIG
    with all KPI definitions, and generates UI_CONFIG so VBA builds the HOME
    sheet automatically when UI_MODE=AUTO.

    Args:
        path:        Path to the .xlsx / .xlsm file to create or update.
        report_name: Human-readable report title (e.g. "Materials Withdrawn").
        tcode:       SAP transaction code (e.g. "MB51").
        date_mode:   CURRENT_WEEK | PREVIOUS_WEEK | MONTH_TO_DATE | YEAR_TO_DATE | CUSTOM
        plant:       SAP plant code (e.g. "1000").
        kpis_json:   JSON array of KPI definitions. Each object supports:
                       kpi_name (required), formula_type (required),
                       column_name (required), target_cell (optional — auto-assigned),
                       filter_column, filter_value, label.
                     formula_type values:
                       SUM, COUNT, COUNT_UNIQUE, AVERAGE, MIN, MAX, FORMULA, CUSTOM
        variant:     SAP selection screen variant name (e.g. "/MY_VARIANT").
        layout:      SAP ALV layout name (e.g. "/MY_LAYOUT").
        use_python:  If True, set UsePython=YES to enable Python cleaning.
        loader_type: BAR (progress bar) or CIRCLE (spinner). Default BAR.

    Example kpis_json:
        [
          {"kpi_name": "Total Qty",    "formula_type": "SUM",          "column_name": "Quantity", "label": "Total Qty"},
          {"kpi_name": "Unique Mats",  "formula_type": "COUNT_UNIQUE",  "column_name": "Material", "label": "Materials"},
          {"kpi_name": "GI Count",     "formula_type": "COUNT",         "column_name": "Movement Type",
           "filter_column": "Movement Type", "filter_value": "261",     "label": "GI Count"}
        ]
    """
    try:
        kpis = json.loads(kpis_json)
        if not isinstance(kpis, list):
            return json.dumps({"error": "kpis_json must be a JSON array"})

        result = wb_ops.build_kpi_dashboard(
            path=path,
            report_name=report_name,
            tcode=tcode,
            date_mode=date_mode,
            plant=plant,
            variant=variant,
            layout=layout,
            kpis=kpis,
            use_python=use_python,
            loader_type=loader_type,
        )
        return json.dumps(result, indent=2)
    except json.JSONDecodeError as exc:
        return json.dumps({"error": f"Invalid kpis_json: {exc}"})
    except Exception as exc:
        return json.dumps({"error": str(exc)})


# ── Tool: apply_theme ─────────────────────────────────────────────────────────
@mcp.tool()
def apply_theme(path: str) -> str:
    """
    Apply Gold/Charcoal theme to all framework sheet tab colours and header rows.

    Args:
        path: Path to the workbook.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        p = Path(path)
        wb = load_workbook(p, keep_vba=p.suffix == ".xlsm")

        for name, headers in wb_ops.SHEET_HEADERS.items():
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            ws.sheet_properties.tabColor = wb_ops.CHARCOAL
            if headers:
                wb_ops._style_header(ws)

        wb.save(p)
        wb.close()
        return json.dumps({"themed": list(wb_ops.SHEET_HEADERS.keys())})
    except Exception as exc:
        return json.dumps({"error": str(exc)})


# ── Tool: inspect_workbook ────────────────────────────────────────────────────
@mcp.tool()
def inspect_workbook(path: str) -> str:
    """
    Return a structural summary of the workbook: sheets, row counts, settings snapshot.

    Args:
        path: Path to the workbook.
    """
    try:
        from openpyxl import load_workbook as lw
        p = Path(path)
        if not p.exists():
            return json.dumps({"error": f"File not found: {path}"})

        wb = lw(p, read_only=True, keep_vba=p.suffix == ".xlsm", data_only=True)
        sheets_info = {
            name: wb[name].max_row for name in wb.sheetnames
        }
        wb.close()

        settings = {}
        try:
            settings = wb_ops.read_settings(p)
        except Exception:
            pass

        kpis = []
        try:
            kpis = wb_ops.list_kpis(p)
        except Exception:
            pass

        return json.dumps({
            "file":     str(p),
            "sheets":   sheets_info,
            "settings": settings,
            "kpi_count": len(kpis),
            "kpis":     kpis,
        }, indent=2)
    except Exception as exc:
        return json.dumps({"error": str(exc)})


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    mcp.run()
