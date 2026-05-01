"""
excel_bridge.py — Read from and write to the .xlsm workbook using openpyxl.

Uses load_workbook with keep_vba=True so macros are never stripped.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import pandas as pd
import openpyxl
from openpyxl import load_workbook

log = logging.getLogger(__name__)


class ExcelBridge:
    """Thin wrapper around openpyxl for safe .xlsm read/write."""

    def __init__(self, workbook_path: Path) -> None:
        self.path = Path(workbook_path)
        if not self.path.exists():
            raise FileNotFoundError(f"Workbook not found: {self.path}")

    # ── Reading ───────────────────────────────────────────────────────────────

    def read_sheet(self, sheet_name: str) -> pd.DataFrame:
        """Return a sheet as a DataFrame (first row = headers)."""
        try:
            wb = load_workbook(self.path, read_only=True, keep_vba=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                log.warning("Sheet not found: %s", sheet_name)
                return pd.DataFrame()
            ws = wb[sheet_name]
            rows = list(ws.values)
            wb.close()
            if not rows:
                return pd.DataFrame()
            return pd.DataFrame(rows[1:], columns=rows[0])
        except Exception as exc:
            log.error("read_sheet(%s) failed: %s", sheet_name, exc)
            return pd.DataFrame()

    def get_setting(self, key: str) -> str:
        """Read a single value from SETTINGS sheet (col A = key, col B = value)."""
        df = self.read_sheet("SETTINGS")
        if df.empty or df.shape[1] < 2:
            return ""
        col_key = df.columns[0]
        col_val = df.columns[1]
        match = df[df[col_key].astype(str).str.strip().str.lower() == key.strip().lower()]
        if match.empty:
            return ""
        return str(match.iloc[0][col_val]).strip()

    # ── Writing ───────────────────────────────────────────────────────────────

    def write_cell(self, sheet_name: str, cell_address: str, value: Any) -> None:
        """Write a single value to a cell, preserving VBA macros."""
        try:
            wb = load_workbook(self.path, keep_vba=True)
            if sheet_name not in wb.sheetnames:
                log.warning("Sheet not found for write: %s", sheet_name)
                wb.close()
                return
            wb[sheet_name][cell_address] = value
            wb.save(self.path)
            wb.close()
            log.info("Wrote %s → %s!%s", value, sheet_name, cell_address)
        except Exception as exc:
            log.error("write_cell(%s!%s) failed: %s", sheet_name, cell_address, exc)

    def write_dataframe(self, sheet_name: str, df: pd.DataFrame,
                        start_row: int = 1, start_col: int = 1,
                        include_header: bool = True) -> None:
        """Write a DataFrame to a sheet starting at (start_row, start_col)."""
        try:
            wb = load_workbook(self.path, keep_vba=True)
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
            ws = wb[sheet_name]

            r = start_row
            if include_header:
                for c, col in enumerate(df.columns, start=start_col):
                    ws.cell(row=r, column=c, value=col)
                r += 1

            for _, row in df.iterrows():
                for c, val in enumerate(row, start=start_col):
                    ws.cell(row=r, column=c, value=val)
                r += 1

            wb.save(self.path)
            wb.close()
            log.info("Wrote %d rows to %s", len(df), sheet_name)
        except Exception as exc:
            log.error("write_dataframe(%s) failed: %s", sheet_name, exc)

    def clear_sheet(self, sheet_name: str) -> None:
        """Delete all data rows in a sheet (keeps the sheet itself)."""
        try:
            wb = load_workbook(self.path, keep_vba=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None
            wb.save(self.path)
            wb.close()
        except Exception as exc:
            log.error("clear_sheet(%s) failed: %s", sheet_name, exc)
